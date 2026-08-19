from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from plan_core.cpm import ScheduleError, compute_schedule
from plan_core.models import Plan, Task

HEADERS = ("задача", "описание", "исполнитель", "длительность", "предшественники")


def _parse_predecessors(raw: object, name_to_id: dict[str, int], row_no: int) -> list[int]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return []
    text = str(raw).strip()
    parts = [p.strip() for p in text.replace(";", ",").split(",") if p.strip()]
    ids: list[int] = []
    for part in parts:
        if part.isdigit():
            ids.append(int(part))
            continue
        key = part.casefold()
        if key not in name_to_id:
            raise ScheduleError(f"Row {row_no}: unknown predecessor '{part}'")
        ids.append(name_to_id[key])
    return ids


def import_plan_from_bytes(
    data: bytes,
    *,
    project_start: date | None = None,
) -> Plan:
    """Parse Excel with Russian headers. Dates come from CPM, not the file."""
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ScheduleError("Excel is empty")

    header = [str(c).strip().casefold() if c is not None else "" for c in rows[0]]
    expected = [h.casefold() for h in HEADERS]
    if header[:5] != expected:
        raise ScheduleError(f"Expected headers {list(HEADERS)}, got {header[:5]}")

    draft: list[dict] = []
    for i, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row[:5]):
            continue
        name = "" if row[0] is None else str(row[0]).strip()
        if not name:
            raise ScheduleError(f"Row {i}: empty task name")
        desc = "" if row[1] is None else str(row[1]).strip()
        assignee = "" if row[2] is None else str(row[2]).strip()
        try:
            duration = int(row[3])
        except (TypeError, ValueError) as exc:
            raise ScheduleError(f"Row {i}: invalid duration") from exc
        if duration < 1:
            raise ScheduleError(f"Row {i}: duration must be >= 1")
        draft.append(
            {
                "row": i,
                "name": name,
                "description": desc,
                "assignee": assignee,
                "duration_days": duration,
                "pred_raw": row[4] if len(row) > 4 else None,
            }
        )

    if not draft:
        raise ScheduleError("No tasks found in Excel")

    name_to_id: dict[str, int] = {}
    for idx, item in enumerate(draft, start=1):
        key = item["name"].casefold()
        if key in name_to_id:
            raise ScheduleError(f"Duplicate task name: {item['name']}")
        name_to_id[key] = idx

    tasks: list[Task] = []
    for idx, item in enumerate(draft, start=1):
        preds = _parse_predecessors(item["pred_raw"], name_to_id, item["row"])
        # Allow numeric refs that match assigned ids
        for p in preds:
            if p < 1 or p > len(draft):
                raise ScheduleError(f"Row {item['row']}: predecessor id {p} out of range")
        tasks.append(
            Task(
                id=idx,
                name=item["name"],
                description=item["description"],
                assignee=item["assignee"],
                duration_days=item["duration_days"],
                predecessor_ids=preds,
            )
        )

    start = project_start or date.today()
    return compute_schedule(Plan(project_start=start, tasks=tasks))


def export_plan_to_bytes(plan: Plan) -> bytes:
    """Write Excel; predecessors as comma-separated task names."""
    by_id = {t.id: t for t in plan.tasks}
    wb = Workbook()
    ws = wb.active
    ws.title = "plan"
    ws.append(list(HEADERS))
    for t in plan.tasks:
        names = []
        for pid in t.predecessor_ids:
            pred = by_id.get(pid)
            names.append(pred.name if pred else str(pid))
        ws.append(
            [
                t.name,
                t.description,
                t.assignee,
                t.duration_days,
                ", ".join(names),
            ]
        )
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
